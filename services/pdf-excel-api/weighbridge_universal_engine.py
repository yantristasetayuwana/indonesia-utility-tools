#!/usr/bin/env python3
"""
Universal Weighbridge PDF -> XLSX engine
V3.2 - flexible metadata/title/total capture.

Features:
- 1 PDF or a directory of PDFs
- no fixed page count
- discovers report/date boundaries from PDF text
- captures title/metadata text instead of throwing it away
- captures report totals/subtotals when recognizable
- preserves all detected transaction rows
- writes a human-readable Weighbridge Data sheet
- writes Report Summary + Diagnostics
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


MONTHS = (
    "January|February|March|April|May|June|July|August|"
    "September|October|November|December"
)

DATE_RE = re.compile(
    rf"\b(\d{{1,2}}\s+(?:{MONTHS})\s+\d{{4}})\b", re.I
)
PAGE_RE = re.compile(r"Page\s*:\s*(\d+)\s+of\s+(\d+)", re.I)

TICKET_RE = re.compile(
    r"\b(?:TLP\d{6,}|GY\d{1,}|TO\d{5,})\b", re.I
)
TIME_RE = re.compile(r"\b\d{1,2}:\d{2}\b")
ROW_RE = re.compile(
    r"^\s*(\d+)\s+"
    r"([A-Za-z]{2,5}\d{1,12})\s+"
    r"(.+?)\s+"
    r"(\d{1,2}:\d{2})\s+"
    r"(\d{1,2}:\d{2})\s+"
    r"(\d+(?:\.\d+)?)\s+"
    r"(\d+(?:\.\d+)?)\s+"
    r"(\d+(?:\.\d+)?)\s+"
    r"(.+)$"
)

HEADER_WORDS = {
    "no", "no.ticket", "truck", "no.truck", "in", "out",
    "gross(ton)", "tarra(ton)", "netto(ton)", "sub",
    "sub leader", "leader",
}

TOTAL_LABEL_RE = re.compile(
    r"\b(?:grand\s+total|total\s+netto|netto\s+total|"
    r"subtotal|sub\s*total|total)\b",
    re.I,
)

# Prefer a number with exactly 3 decimals because weighbridge totals
# normally use the same precision as Netto(Ton).
WEIGHT3_RE = re.compile(r"\b\d+\.\d{3}\b")


def clean(s) -> str:
    return " ".join(str(s or "").replace("\x00", " ").split()).strip()


def norm_id(s) -> str:
    return re.sub(r"\s+", "", clean(s)).upper()


def dec(s):
    try:
        return Decimal(clean(s).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def split_candidate_lines(text):
    return [clean(x) for x in text.splitlines() if clean(x)]


def extract_pages(pdf: Path):
    reader = PdfReader(str(pdf))
    pages = []
    for pno, page in enumerate(reader.pages, 1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            text = ""
            print(f"WARN page {pno}: extraction failed: {exc}")
        pages.append((pno, text))
    return pages


def discover_title(lines):
    """
    Capture report title/heading without assuming an exact title string.
    The first heading containing WEIGHBRIDGE / DAILY REPORT is preferred.
    """
    candidates = []

    for i, line in enumerate(lines[:25]):
        low = line.lower()
        if "weighbridge" in low or "daily report" in low:
            candidates.append(line)

            # Some PDFs split the title over 2 lines.
            if i + 1 < len(lines):
                nxt = lines[i + 1]
                nlow = nxt.lower()
                if (
                    nxt
                    and not DATE_RE.search(nxt)
                    and "page :" not in nlow
                    and not any(h in nlow for h in ("no.ticket", "gross(ton)", "tarra(ton)"))
                ):
                    if len(nxt) <= 120:
                        candidates.append(nxt)
            break

    if candidates:
        return " | ".join(dict.fromkeys(candidates))

    # Fallback: first substantial non-table heading.
    for line in lines[:20]:
        low = line.lower()
        if (
            len(line) >= 5
            and not DATE_RE.search(line)
            and not PAGE_RE.search(line)
            and not ROW_RE.match(line)
            and not TICKET_RE.search(line)
            and not any(h in low for h in HEADER_WORDS)
        ):
            return line

    return ""


def discover_total(lines, row_indexes=None):
    """
    Find a report subtotal/total from text around the end of the report.
    This is deliberately label-based, not page-number-based.
    """
    row_indexes = row_indexes or []
    start = (max(row_indexes) + 1) if row_indexes else max(0, len(lines) - 15)

    # Search after the last transaction first.
    search_ranges = [
        range(start, len(lines)),
        range(max(0, len(lines) - 20), len(lines)),
    ]

    seen = set()
    for rng in search_ranges:
        for i in rng:
            if i in seen:
                continue
            seen.add(i)

            line = lines[i]
            if not TOTAL_LABEL_RE.search(line):
                continue

            nums = WEIGHT3_RE.findall(line)
            if not nums and i + 1 < len(lines):
                nums = WEIGHT3_RE.findall(lines[i + 1])

            if nums:
                # The last number on a total line is usually the displayed total.
                return Decimal(nums[-1])

    return None


def discover_rows(text):
    """
    Flexible transaction-row parser.
    Finds ticket/time/weight structure even when supplier/leader text wraps.
    """
    lines = split_candidate_lines(text)
    rows = []
    row_indexes = []
    i = 0

    while i < len(lines):
        line = lines[i]
        m = ROW_RE.match(line)

        if m:
            no, ticket, truck, tin, tout, gross, tarra, netto, tail = m.groups()
            rows.append([
                no, clean(ticket), clean(truck), tin, tout,
                gross, tarra, netto, clean(tail)
            ])
            row_indexes.append(i)
            i += 1
            continue

        tm = TICKET_RE.search(line)
        if tm:
            window_lines = lines[i:i + 4]
            window = " ".join(window_lines)
            times = TIME_RE.findall(window)
            nums = WEIGHT3_RE.findall(window)

            if len(times) >= 2 and len(nums) >= 3:
                before = line[:tm.start()].strip()
                no_m = re.search(r"^\s*(\d+)\b", before)
                no = no_m.group(1) if no_m else ""
                ticket = tm.group(0)

                truck_m = re.search(
                    r"(?:^|\s)([A-Z]{1,3}\s*\d{2,5}[A-Z]{0,4})\s+"
                    + re.escape(times[0]),
                    window,
                    re.I,
                )
                truck = truck_m.group(1) if truck_m else ""

                rows.append([
                    no, clean(ticket), clean(truck),
                    times[0], times[1], nums[0], nums[1], nums[2],
                    clean(line[tm.end():])
                ])
                row_indexes.append(i)
                i += 1
                continue

        i += 1

    # Remove accidental duplicate candidates.
    out = []
    seen = set()
    kept_indexes = []

    for idx, r in zip(row_indexes, rows):
        key = norm_id(r[1])
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(r)
        kept_indexes.append(idx)

    return out, kept_indexes


def improve_rows(rows):
    """Split final supplier/leader tail into Sub + Leader."""
    result = []

    for r in rows:
        tail = clean(r[8])
        parts = re.split(r"\s{2,}", tail)

        if len(parts) >= 2:
            sub, leader = parts[0], parts[-1]
        else:
            words = tail.split()
            half = len(words) // 2
            if half and " ".join(words[:half]) == " ".join(words[half:]):
                sub = leader = " ".join(words[:half])
            else:
                sub = leader = tail

        result.append(r[:8] + [sub, leader])

    return result


def discover_reports(pages):
    """
    Discover report boundaries without fixed page numbers.

    A report starts when a page contains a date plus a report marker.
    Metadata/title/total are captured and retained for XLSX output.
    """
    reports = []

    for pno, text in pages:
        lines = split_candidate_lines(text)
        dates = list(DATE_RE.finditer(text))
        if not dates:
            continue

        low = text.lower()
        if "weighbridge" not in low and "daily report" not in low:
            continue

        date = dates[0].group(1)
        pm = PAGE_RE.search(text)
        page_total = int(pm.group(2)) if pm else None

        rows, row_indexes = discover_rows(text)
        title = discover_title(lines)
        total = discover_total(lines, row_indexes)

        reports.append({
            "page": pno,
            "date": date,
            "page_total": page_total,
            "title": title,
            "total": total,
            "rows": rows,
        })

    return reports


def validate(rows):
    errors = []
    seen = set()

    for idx, r in enumerate(rows, 2):
        ticket = norm_id(r[1])

        if not ticket:
            errors.append((idx, "EMPTY_TICKET"))
        if ticket in seen:
            errors.append((idx, "DUPLICATE_TICKET", ticket))
        seen.add(ticket)

        g, t, n = dec(r[5]), dec(r[6]), dec(r[7])

        if g is None or t is None or n is None:
            errors.append((idx, "BAD_WEIGHT", r[1]))
        elif abs((g - t) - n) > Decimal("0.005"):
            errors.append((idx, "MATH", r[1], str(g), str(t), str(n)))

        if not clean(r[8]):
            errors.append((idx, "EMPTY_SUB", r[1]))
        if not clean(r[9]):
            errors.append((idx, "EMPTY_LEADER", r[1]))

    return errors


def style_title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 24


def style_meta(ws, row, label, value):
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


def style_headers(ws, row):
    headers = [
        "No", "No.Ticket", "Truck", "In", "Out",
        "Gross(Ton)", "Tarra(Ton)", "Netto(Ton)", "Sub", "Leader"
    ]
    for col, value in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:J{row}"


def write_xlsx(rows, reports, pdf, outdir):
    outdir.mkdir(parents=True, exist_ok=True)
    out = outdir / f"hasil_{pdf.stem}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Weighbridge Data"

    # Keep title and metadata INSIDE the main Excel sheet.
    current_row = 1
    pos = 0

    for report_no, rep in enumerate(reports, 1):
        title = rep.get("title") or "WEIGHBRIDGE REPORT"

        style_title(ws, current_row, title)
        current_row += 1

        style_meta(ws, current_row, "Report Date", rep.get("date", ""))
        current_row += 1

        style_meta(
            ws, current_row, "Source Page",
            f"{rep.get('page')} / {rep.get('page_total') or '?'}"
        )
        current_row += 1

        style_headers(ws, current_row)
        current_row += 1

        chunk = rows[pos:pos + len(rep["rows"])]

        for r in chunk:
            for col, value in enumerate(r, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

        excel_total = sum((dec(r[7]) or Decimal(0)) for r in chunk)
        pdf_total = rep.get("total")

        # Always show Excel calculated total.
        ws.cell(row=current_row, column=7, value="REPORT TOTAL")
        ws.cell(row=current_row, column=7).font = Font(bold=True)
        ws.cell(row=current_row, column=8, value=float(excel_total))
        ws.cell(row=current_row, column=8).font = Font(bold=True)

        if pdf_total is not None:
            ws.cell(row=current_row, column=9, value="PDF TOTAL")
            ws.cell(row=current_row, column=9).font = Font(bold=True)
            ws.cell(row=current_row, column=10, value=float(pdf_total))
            ws.cell(row=current_row, column=10).font = Font(bold=True)

        current_row += 2
        pos += len(chunk)

    # Column widths.
    widths = {
        "A": 8, "B": 18, "C": 16, "D": 10, "E": 10,
        "F": 14, "G": 14, "H": 14, "I": 34, "J": 34
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Summary sheet.
    rs = wb.create_sheet("Report Summary")
    rs.append([
        "Report", "Page", "Date", "RIT",
        "PDF Total Netto", "Excel Total Netto", "Diff", "Status", "Title"
    ])

    pos = 0
    for i, rep in enumerate(reports, 1):
        chunk = rows[pos:pos + len(rep["rows"])]
        excel_total = sum((dec(r[7]) or Decimal(0)) for r in chunk)
        pdf_total = rep.get("total")

        if pdf_total is not None:
            diff = excel_total - pdf_total
            status = "PASS" if abs(diff) <= Decimal("0.005") else "CHECK"
            diff_value = float(diff)
            pdf_value = float(pdf_total)
        else:
            diff_value = ""
            pdf_value = ""
            status = "NO PDF TOTAL"

        rs.append([
            i,
            rep.get("page"),
            rep.get("date"),
            len(chunk),
            pdf_value,
            float(excel_total),
            diff_value,
            status,
            rep.get("title", ""),
        ])
        pos += len(chunk)

    for cell in rs[1]:
        cell.font = Font(bold=True)
    rs.freeze_panes = "A2"
    rs.column_dimensions["I"].width = 60

    # Diagnostics.
    diag = wb.create_sheet("Diagnostics")
    diag.append(["Metric", "Value"])
    diag.append(["Source PDF", pdf.name])
    diag.append(["Pages", sum(1 for _ in extract_pages(pdf))])
    diag.append(["Reports detected", len(reports)])
    diag.append(["Rows extracted", len(rows)])
    diag.append(["Unique tickets", len({norm_id(r[1]) for r in rows})])
    diag.append(["Rows consumed", pos])
    diag.append(["Rows remaining", len(rows) - pos])
    diag.append([
        "Reports with title",
        sum(1 for r in reports if r.get("title"))
    ])
    diag.append([
        "Reports with PDF total",
        sum(1 for r in reports if r.get("total") is not None)
    ])

    for cell in diag[1]:
        cell.font = Font(bold=True)
    diag.column_dimensions["A"].width = 30
    diag.column_dimensions["B"].width = 70

    wb.save(out)
    return out


def process(pdf: Path, outdir: Path):
    print("=" * 80)
    print(f"UNIVERSAL WEIGHBRIDGE ENGINE V3.2 | {pdf.name}")
    print("=" * 80)

    pages = extract_pages(pdf)
    reports = discover_reports(pages)

    all_rows = []

    for rep in reports:
        rep["rows"] = improve_rows(rep["rows"])
        all_rows.extend(rep["rows"])

    errors = validate(all_rows)
    tickets = [norm_id(r[1]) for r in all_rows]

    print(f"PAGES              : {len(pages)}")
    print(f"REPORTS DETECTED   : {len(reports)}")
    print(f"ROWS EXTRACTED     : {len(all_rows)}")
    print(f"UNIQUE TICKETS     : {len(set(tickets))}")
    print(f"DUPLICATE TICKETS  : {len(tickets) - len(set(tickets))}")
    print(f"VALIDATION ERRORS  : {len(errors)}")

    for i, rep in enumerate(reports, 1):
        print(
            f"REPORT {i:02d} | PAGE={rep['page']} | "
            f"DATE={rep['date']} | ROWS={len(rep['rows'])} | "
            f"TITLE={'YES' if rep.get('title') else 'NO'} | "
            f"PDF_TOTAL={'YES' if rep.get('total') is not None else 'NO'}"
        )

    if errors:
        for e in errors[:30]:
            print(" -", e)

    out = write_xlsx(all_rows, reports, pdf, outdir)
    print(f"OUTPUT             : {out}")

    hard = [
        e for e in errors
        if e[1] in {"EMPTY_TICKET", "DUPLICATE_TICKET", "BAD_WEIGHT", "MATH"}
    ]

    if not reports:
        print("GATE               : FAIL (no reports detected)")
        return 2

    if not all_rows:
        print("GATE               : FAIL (no rows detected)")
        return 2

    if hard:
        print("GATE               : CHECK REQUIRED")
        return 1

    print("GATE               : PASS")
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="PDF file or directory containing PDFs")
    ap.add_argument("--out", default="output_weighbridge")
    args = ap.parse_args()

    src = Path(args.input)
    outdir = Path(args.out)

    if src.is_file():
        return process(src, outdir)

    if src.is_dir():
        pdfs = sorted(src.glob("*.pdf"))
        if not pdfs:
            print("No PDF files found.")
            return 2

        codes = [process(pdf, outdir) for pdf in pdfs]
        return (
            1 if any(c == 1 for c in codes)
            else (2 if any(c == 2 for c in codes) else 0)
        )

    print(f"Input not found: {src}")
    return 2


def convert_pdf_to_excel(pdf, outdir):
    """API-safe alias: convert one PDF and return the generated XLSX Path."""
    pdf = Path(pdf)
    outdir = Path(outdir)
    process(pdf, outdir)
    output = outdir / f"hasil_{pdf.stem}.xlsx"
    if not output.exists():
        candidates = sorted(outdir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
        if not candidates:
            raise FileNotFoundError(f"No XLSX generated for {pdf.name}")
        output = candidates[-1]
    return output


if __name__ == "__main__":
    sys.exit(main())
