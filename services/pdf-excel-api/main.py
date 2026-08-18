import io, os, re, time, json, math, asyncio
from typing import Any, Dict, List, Tuple
import httpx
from fastapi import FastAPI, UploadFile, File, Header, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

app = FastAPI(title="Indonesia Utility Tools PDF Excel V3")

AZURE_ENDPOINT = os.getenv("AZURE_DI_ENDPOINT","").rstrip("/")
AZURE_KEY = os.getenv("AZURE_DI_KEY","")
SERVICE_TOKEN = os.getenv("PDF_EXCEL_API_TOKEN","")
MAX_SECONDS = int(os.getenv("AZURE_MAX_SECONDS","240"))

def clean(v):
    return re.sub(r"\s+"," ",str(v or "")).strip()

def azure_headers():
    return {"Ocp-Apim-Subscription-Key": AZURE_KEY}

async def azure_analyze(pdf: bytes):
    if not AZURE_ENDPOINT or not AZURE_KEY:
        raise RuntimeError("Azure Document Intelligence belum dikonfigurasi.")
    url=f"{AZURE_ENDPOINT}/documentintelligence/documentModels/prebuilt-layout:analyze?api-version=2024-11-30"
    async with httpx.AsyncClient(timeout=90) as c:
        r=await c.post(url,headers={**azure_headers(),"Content-Type":"application/pdf"},content=pdf)
        if r.status_code != 202:
            raise RuntimeError(f"Azure analyze gagal ({r.status_code}): {r.text[:700]}")
        op=r.headers.get("operation-location")
        if not op: raise RuntimeError("Azure tidak mengembalikan operation-location.")
        deadline=time.time()+MAX_SECONDS
        while time.time()<deadline:
            await asyncio.sleep(1.5)
            q=await c.get(op,headers=azure_headers())
            if q.status_code>=400: raise RuntimeError(f"Azure polling gagal ({q.status_code}): {q.text[:700]}")
            data=q.json()
            st=str(data.get("status","")).lower()
            if st=="succeeded": return data
            if st=="failed": raise RuntimeError("Azure gagal menganalisis PDF.")
        raise RuntimeError("Waktu analisis Azure habis.")

def table_matrix(table):
    nrows=int(table.get("rowCount",0)); ncols=int(table.get("columnCount",0))
    mat=[["" for _ in range(ncols)] for _ in range(nrows)]
    for cell in table.get("cells",[]):
        r=int(cell.get("rowIndex",0)); c=int(cell.get("columnIndex",0))
        if r>=nrows or c>=ncols: continue
        txt=clean(cell.get("content",""))
        mat[r][c]=txt
    return mat

def choose_tables(result):
    tables=result.get("analyzeResult",{}).get("tables",[]) or []
    # Prefer the largest table; many PDFs have one logical table repeated across pages.
    return sorted(tables,key=lambda t:(int(t.get("rowCount",0))*int(t.get("columnCount",0)), int(t.get("rowCount",0))), reverse=True)

def merge_tables(tables):
    if not tables: return [], []
    chosen=[]
    for t in tables:
        m=table_matrix(t)
        if not m: continue
        chosen.append(m)
    if not chosen: return [], []
    # Normalize width and identify header from the first non-empty row.
    width=max(len(r) for m in chosen for r in m)
    normalized=[[clean(x) for x in r]+[""]*(width-len(r)) for m in chosen for r in m]
    # Remove exact repeated headers, blank rows, and page-number-only rows.
    header=normalized[0]
    out=[header]
    for r in normalized[1:]:
        if not any(r): continue
        if [clean(x).lower() for x in r]==[clean(x).lower() for x in header]: continue
        if re.fullmatch(r"page\s*\d+(\s*of\s*\d+)?", " ".join(r), re.I): continue
        out.append(r)
    return header,out[1:]

def extract_layout(result):
    ar=result.get("analyzeResult",{})
    pages=ar.get("pages",[]) or []
    rows=[]
    for p in pages:
        pn=p.get("pageNumber")
        for line in p.get("lines",[]) or []:
            rows.append([pn,"CONTENT",clean(line.get("content",""))])
    # Azure may expose pageHeader/pageFooter in newer API shapes.
    for p in pages:
        pn=p.get("pageNumber")
        for x in p.get("pageHeader",[]) or []: rows.append([pn,"HEADER",clean(x.get("content",x) if isinstance(x,dict) else x)])
        for x in p.get("pageFooter",[]) or []: rows.append([pn,"FOOTER",clean(x.get("content",x) if isinstance(x,dict) else x)])
    return rows

def validate(headers, rows, result):
    warnings=[]
    if not headers: warnings.append("Header tabel tidak terdeteksi.")
    if not rows: warnings.append("Tidak ada data row terdeteksi.")
    widths={len(r) for r in rows}
    if widths and len(widths)>1: warnings.append("Lebar row tidak seragam; telah dinormalisasi.")
    # Detect a numeric No column and gaps.
    seq=[]
    for r in rows:
        if r and re.fullmatch(r"\d{1,6}",clean(r[0])): seq.append(int(r[0]))
    if len(seq)>=3:
        gaps=[f"{a}->{b}" for a,b in zip(seq,seq[1:]) if b!=a+1]
        if gaps: warnings.append("Urutan nomor memiliki gap: "+", ".join(gaps[:12]))
    return warnings

def build_workbook(headers, rows, layout, diagnostics):
    wb=Workbook()
    ws=wb.active; ws.title="Converted Table"
    ws.append(headers)
    for r in rows: ws.append(r)
    thin=Side(style="thin")
    for cell in ws[1]:
        cell.font=Font(bold=True)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=Border(top=thin,bottom=thin,left=thin,right=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment=Alignment(vertical="top",wrap_text=True)
            c.border=Border(bottom=thin)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col in range(1,len(headers)+1):
        maxlen=max([len(clean(ws.cell(r,col).value)) for r in range(1,ws.max_row+1)]+[8])
        ws.column_dimensions[get_column_letter(col)].width=min(50,max(10,maxlen+2))
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.page_margins=PageMargins(left=.25,right=.25,top=.5,bottom=.5,header=.2,footer=.2)

    lay=wb.create_sheet("PDF Layout"); lay.append(["Page","Role","Text"])
    for r in layout: lay.append(r)
    for c in lay[1]: c.font=Font(bold=True)
    lay.freeze_panes="A2"
    lay.column_dimensions["A"].width=10; lay.column_dimensions["B"].width=14; lay.column_dimensions["C"].width=100
    for row in lay.iter_rows(): row[2].alignment=Alignment(wrap_text=True,vertical="top")

    raw=wb.create_sheet("Raw Text")
    raw.append(["Page","Text"])
    for r in layout:
        if r[1]=="CONTENT": raw.append([r[0],r[2]])
    raw.freeze_panes="A2"; raw.column_dimensions["B"].width=100

    dg=wb.create_sheet("Diagnostics")
    dg.append(["Metric","Value"])
    for k,v in diagnostics.items(): dg.append([k,str(v)])
    for c in dg[1]: c.font=Font(bold=True)
    dg.column_dimensions["A"].width=28; dg.column_dimensions["B"].width=100
    for row in dg.iter_rows(): row[1].alignment=Alignment(wrap_text=True,vertical="top")
    return wb

@app.post("/convert")
async def convert(file: UploadFile=File(...), authorization: str|None=Header(default=None)):
    if SERVICE_TOKEN and authorization != f"Bearer {SERVICE_TOKEN}":
        raise HTTPException(401,"Unauthorized")
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400,"Hanya PDF yang didukung.")
    pdf=await file.read()
    if len(pdf)>50*1024*1024: raise HTTPException(413,"PDF maksimum 50 MB.")
    try:
        result=await azure_analyze(pdf)
        tables=choose_tables(result)
        headers,rows=merge_tables(tables)
        if not headers or not rows: raise RuntimeError("Azure tidak menemukan tabel yang dapat dikonversi.")
        warnings=validate(headers,rows,result)
        pages=len(result.get("analyzeResult",{}).get("pages",[]) or [])
        layout=extract_layout(result)
        diagnostics={
            "Engine":"Azure Document Intelligence + coordinate table reconstruction + openpyxl",
            "Pages":pages,"Columns":len(headers),"Rows":len(rows),
            "First row":rows[0][0] if rows else "","Last row":rows[-1][0] if rows else "",
            "Warnings":" | ".join(warnings) if warnings else "PASS",
            "Source filename":file.filename
        }
        if warnings:
            # We still produce the workbook, but the warning is visible in Diagnostics.
            diagnostics["Validation"]="WARNING"
        else: diagnostics["Validation"]="PASS"
        wb=build_workbook(headers,rows,layout,diagnostics)
        bio=io.BytesIO(); wb.save(bio); bio.seek(0)
        outname=re.sub(r"\.pdf$","",file.filename,flags=re.I)+"-converted.xlsx"
        headers_out={
            "Content-Type":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition":f'attachment; filename="{outname}"',
            "X-Converted-Rows":str(len(rows)),
            "X-Converted-Cols":str(len(headers)),
            "X-Converted-Pages":str(pages),
            "X-Validation":"PASS" if not warnings else "WARNING",
            "Cache-Control":"no-store"
        }
        return StreamingResponse(bio,media_type=headers_out["Content-Type"],headers=headers_out)
    except Exception as e:
        return JSONResponse({"ok":False,"error":"PDF → Excel conversion gagal.","detail":str(e)[:1500]},status_code=502)
